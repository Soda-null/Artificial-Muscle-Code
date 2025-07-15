# ==============================================================================
# final_controller.py - v8 (手动确认并实时显示压力版)
# 核心改进：在等待用户确认期间，实时显示当前压力值，以辅助调节。
# ==============================================================================

import serial
import threading
import time
import queue
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import statistics

# ==============================================================================
# --- 1. 配置区 ---
# ==============================================================================
SERIAL_PORT = 'COM3'
BAUD_RATE = 9600

# --- 目标压力提示列表 ---
TARGET_PRESSURES = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6]

# --- 文件路径配置 ---
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
if not os.path.exists(desktop_path):
    desktop_path = "."
OUTPUT_FILENAME = os.path.join(desktop_path, f"structured_experiment_{int(time.time())}.xlsx")

# ==============================================================================
# --- 2. 全局变量 ---
# ==============================================================================
stop_all_threads = threading.Event()
data_queue = queue.Queue()

# ==============================================================================
# --- 3. 功能函数定义 ---
# ==============================================================================

def serial_reader_thread(ser):
    """后台线程：持续从串口读取数据并放入队列。"""
    while not stop_all_threads.is_set():
        try:
            line = ser.readline().decode('utf-8').strip()
            if line:
                parts = line.split(',')
                if len(parts) == 3:
                    data_tuple = tuple(map(float, parts))
                    if data_queue.qsize() > 100:
                        data_queue.get()
                    data_queue.put(data_tuple)
        except (serial.SerialException, OSError, TypeError):
            print("\n串口错误，读取线程退出。")
            stop_all_threads.set()
            break
        except (ValueError, IndexError, UnicodeDecodeError):
            continue

def perform_offset_calibration():
    """步骤一：进行初始长度的校准。"""
    print("\n--- 步骤 1: 长度校准 ---")
    real_initial_length = 0.0
    while real_initial_length <= 0:
        try:
            length_input = input("请输入您测量的物体真实初始长度 (mm)，然后按 [Enter]: ")
            real_initial_length = float(length_input)
        except ValueError:
            print("输入无效。")
    input(f"\n已设置真实初始长度为 {real_initial_length:.2f} mm。\n请将物体放置在该初始位置，然后按 [Enter] 键开始测量传感器读数...")
    while not data_queue.empty(): data_queue.get()
    calibration_duration = 5
    collected_distances = []
    print("正在测量传感器读数...")
    start_time = time.time()
    while time.time() - start_time < calibration_duration:
        try:
            _, distance, _ = data_queue.get(timeout=2.0)
            collected_distances.append(distance)
            time_left = calibration_duration - (time.time() - start_time)
            print(f"\r采集中... 剩余 {time_left:.1f} 秒", end="")
        except queue.Empty:
            print("\n警告：校准期间未收到数据。"); break
    if not collected_distances:
        print("\n错误：校准失败，未收集到任何数据。"); return None, None
    sensor_measured_distance = statistics.mean(collected_distances)
    calibration_offset = real_initial_length - sensor_measured_distance
    print(f"\n\n校准完成！传感器在初始位置的平均读数为: {sensor_measured_distance:.2f} mm")
    print(f"计算出的校准偏移量为: {calibration_offset:.2f} mm")
    return real_initial_length, calibration_offset

# ==================== 这是修改后的核心函数 (带实时压力显示) ====================
def detect_and_lock_pressure():
    """
    (新) 半自动模式 V2：
    1. 用户手动输入要测量的目标压力。
    2. 在等待用户按Enter期间，实时刷新显示当前压力值。
    3. 用户调节好气压后，按Enter键确认，程序直接开始采集。
    """
    print("\n" + "="*50)
    print("--- 步骤 2: 压力目标设置 ---")
    
    locked_target = None
    while not locked_target:
        prompt_str = "请输入本次要测量的目标压力档位 (推荐: " + ", ".join(map(str, TARGET_PRESSURES)) + "),"
        user_input = input(f"{prompt_str}\n或输入 'q' 退出并保存: ")

        if user_input.lower() == 'q':
            return None

        try:
            target_pressure = float(user_input)
            if target_pressure not in TARGET_PRESSURES:
                print(f"警告：输入值 {target_pressure} 不在推荐列表中，但仍将继续。")
            locked_target = target_pressure
        except ValueError:
            print("输入无效，请输入一个数字。")

    # --- 实时压力显示的核心逻辑 ---
    stop_pressure_display = threading.Event()

    def pressure_display_thread():
        # 在新线程中刷新压力
        while not stop_pressure_display.is_set():
            try:
                # 只看不取，或者快速取用
                _, _, current_pressure = data_queue.get(timeout=0.1)
                # 使用 \r 在同一行更新
                print(f"\r请调节气压至 {locked_target} MPa 附近... 当前压力: {current_pressure:.3f} MPa", end="")
            except queue.Empty:
                # 队列为空时，短暂休眠，避免CPU空转
                time.sleep(0.05)
    
    print(f"\n已设定目标压力为 {locked_target} MPa。")
    # 清空旧数据，准备开始显示实时压力
    while not data_queue.empty(): data_queue.get()
    
    # 启动压力显示线程
    display_thread = threading.Thread(target=pressure_display_thread)
    display_thread.daemon = True # 设置为守护线程，主程序退出时它也退出
    display_thread.start()

    # 主线程在这里等待用户按Enter
    input("\n按 [Enter] 键确认并开始数据采集...")
    
    # 用户按了Enter，停止压力显示线程
    stop_pressure_display.set()
    display_thread.join(timeout=0.5) # 等待线程结束

    # 结束后打印一个换行，让后续输出在新的一行开始
    print() 

    # 再次清空队列，确保数据采集从一个干净的状态开始
    while not data_queue.empty(): data_queue.get()
        
    return locked_target
# ===========================================================================

def collect_data_for_session(real_initial_length, calibration_offset):
    """步骤三：在设定的压力下采集数据，直到用户按Enter。"""
    session_data = []
    print("\n--- 步骤 3: 数据采集 ---")
    print(">>> 数据采集中... 按 [Enter] 结束当前测量。")
    
    stop_collecting_flag = threading.Event()
    def wait_for_enter():
        input()
        stop_collecting_flag.set()

    input_thread = threading.Thread(target=wait_for_enter)
    input_thread.start()
    
    while not stop_collecting_flag.is_set():
        try:
            force, sensor_dist, pressure = data_queue.get(timeout=2.0)
            calibrated_dist = sensor_dist + calibration_offset
            shrinkage = (-(calibrated_dist - real_initial_length) / real_initial_length) * 100.0
            
            force_str = f"力: {force:6.2f} N"
            shrink_str = f"收缩率: {shrinkage:6.2f} %"
            pressure_str = f"实时压力: {pressure:5.3f} MPa"
            print(f"\r{force_str}  |  {shrink_str}  |  {pressure_str}   ", end="")
            
            session_data.append((force, shrinkage))
        except queue.Empty:
            time.sleep(0.01)
            continue
            
    print("\n--- 采集结束 ---")
    input_thread.join()
    return session_data

def save_to_excel(filename, data):
    """将所有采集到的数据分列保存到Excel文件中。"""
    if not data:
        print("没有采集到任何数据，无需保存。")
        return
    print(f"\n正在保存数据到 {filename} ...")
    wb = Workbook()
    ws = wb.active
    ws.title = "实验数据"
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    pressures_keys = sorted(data.keys())
    ws.cell(row=1, column=1, value="实验数据记录").font = Font(bold=True, size=16)
    header_row = 2
    
    for col_idx, pressure_key in enumerate(pressures_keys):
        start_col = col_idx * 2 + 1
        title_cell = ws.cell(row=header_row, column=start_col, value=f"气压: {pressure_key} MPa")
        title_cell.font = header_font
        title_cell.alignment = center_alignment
        ws.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=start_col + 1)
        ws.cell(row=header_row + 1, column=start_col, value="力 (N)").font = header_font
        ws.cell(row=header_row + 1, column=start_col + 1, value="收缩率 (%)").font = header_font
        records = data[pressure_key]
        for row_idx, (force, shrinkage) in enumerate(records):
            ws.cell(row=header_row + 2 + row_idx, column=start_col, value=force)
            ws.cell(row=header_row + 2 + row_idx, column=start_col + 1, value=shrinkage)
            
    for col_idx in range(1, len(pressures_keys) * 2 + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
        
    try:
        wb.save(filename)
        print(f"成功！数据已保存到 {filename}")
    except Exception as e:
        print(f"保存文件失败: {e}")

# ==============================================================================
# --- 4. 主程序入口 ---
# ==============================================================================
def main():
    reader_thread = None
    all_sessions_data = {}
    try:
        ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=5)
        print(f"成功连接到串口 {SERIAL_PORT}。")
        print("正在等待Arduino的“准备就绪”信号...")
        handshake_received = False
        while not handshake_received:
            line = ser.readline().decode('utf-8').strip()
            if line == "Arduino is Ready":
                handshake_received = True
                print("握手成功！Arduino已准备就绪。")
            elif not line:
                print("错误：等待握手信号超时。程序退出。")
                return
        
        ser.reset_input_buffer()
        reader_thread = threading.Thread(target=serial_reader_thread, args=(ser,))
        reader_thread.daemon = True
        reader_thread.start()
        print("后台数据接收线程已启动。")

        real_initial_length, calibration_offset = perform_offset_calibration()
        if real_initial_length is None:
            return

        while not stop_all_threads.is_set():
            locked_pressure = detect_and_lock_pressure()
            if locked_pressure is None: 
                break
            
            session_data = collect_data_for_session(real_initial_length, calibration_offset)
            if session_data:
                all_sessions_data[locked_pressure] = session_data
            
    except serial.SerialException as e:
        print(f"\n错误：无法打开串口 {SERIAL_PORT}。{e}")
        print("请检查：1. 设备是否已连接。 2. 串口号是否正确。 3. 其他程序是否占用了该串口。")
    except KeyboardInterrupt:
        print("\n\n检测到 Ctrl+C，程序将退出。")
    finally:
        stop_all_threads.set()
        if reader_thread and reader_thread.is_alive():
            reader_thread.join(timeout=1.0)
        
        save_to_excel(OUTPUT_FILENAME, all_sessions_data)
        print("\n程序结束。")

if __name__ == "__main__":
    main()